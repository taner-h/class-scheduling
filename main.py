import json
import time
import numpy as np
import random
import copy
import pickle
import openpyxl
import os
from collections import Counter


class Course:
    def __init__(self, id, name, code, department, year, cannotCollideWith):
        self.id = id
        self.name = name
        self.code = code
        self.department = department
        self.year = year
        self.cannotCollideWith = cannotCollideWith

    def print(self):
        print(f"{self.id}: {self.name} ({self.code}), {self.department}/{self.year}")


class Teacher:
    def __init__(self, id, firstName, lastName, unavailable):
        self.id = id
        self.firstName = firstName
        self.lastName = lastName
        self.unavailable = unavailable

    def print(self):
        print(f"{self.id}: {self.firstName} {self.lastName}")


class Session:
    def __init__(self, id, course, teacher, length, isLab=False, suffix=None, isFixed=False, day=None, hour=None):
        self.id = id
        self.course = course
        self.teacher = teacher
        self.length = length

        self.name = self.course.name if not isLab else self.course.name + ' - ' + suffix

        self.isFixed = isFixed
        self.day = day
        self.hour = hour

    def print(self):
        print(
            f"{self.id}: {self.name} ({self.length}), {self.teacher.firstName} {self.teacher.lastName}")

    def printSchedule(self):
        print(f"{self.id}: {self.name}, {getDayName(self.day)} {self.hour}.00 - {self.hour + self.length}.00 ({getSemesterShortName((4 * self.course.department + self.course.year) - 1)})")


class Schedule:
    def __init__(self, state):
        self.state = state

        self.semesters = self.filterSemesters()
        self.teacherSessions = self.filterByTeachers()

        self.semesterCollisions = self.calculateSemesterCollisions()
        self.teacherCollisions = self.calculateTeacherCollisions()
        self.multiTeacherCollisions = self.calculateMultiTeacherSessionCollisions()
        self.teacherAvailabilityViolations = self.calculateTeacherAvailabilityViolations()
        self.cannotCollideViolations = self.calculateCannotCollideViolations()
        results = self.calculateConstraints()

        self.breakHourViolations = results[0]
        self.departmentMeetingViolations = results[1]
        self.fridayBreakViolations = results[2]
        self.languageSessionViolations = results[3]
        self.freeDays = results[4]
        self.singleSessionDays = results[5]
        self.multipleCourseSessions = results[6]
        self.slotSpan = results[7]
        self.availableSlots = results[8]
        self.allSlotsUsedDays = self.calculateAllSlotsUsedDays()

        isValid, fitness = self.calculateFitness()
        self.fitness = fitness
        self.isValid = isValid
        self.hasAllSessions = self.calculateHasAllSessions()

    def filterSemesters(self):
        semesters = []
        for department in range(2):
            for year in range(1, 5):
                semester = list(filter(lambda session: session.course.department ==
                                       department and session.course.year == year, self.state))
                semesters.append(semester)

        return semesters

    def filterByTeachers(self):
        filtered = []
        for teacher in teachers:
            teacherSessions = list(
                filter(lambda session: session.teacher.id == teacher.id, self.state))
            filtered.append(teacherSessions)
        return filtered

    def print(self):
        for index, semester in enumerate(self.semesters):
            print(
                f"\n\n----- {getSemesterName(index // 4, index % 4 + 1)} -----\n\n")
            for day in range(5):
                semesterCopy = copy.deepcopy(semester)
                sessionsOfDay = list(filter(
                    lambda session: session.day == day, semesterCopy))
                ordered = sorted(
                    sessionsOfDay, key=lambda session: session.hour)
                print(f'\n{getDayName(day)}\n')
                for session in ordered:
                    print(
                        f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name}')
                if day in [1, 2]:
                    print('16.00 - 18.00 - Yabancı Dil')

    def printTeacherSessions(self):
        for sessionsOfTeacher in self.teacherSessions:
            if sessionsOfTeacher[0].teacher.firstName != "":
                print(
                    f'\n\n----- {sessionsOfTeacher[0].teacher.firstName} {sessionsOfTeacher[0].teacher.lastName} -----\n')
                if sessionsOfTeacher[0].teacher.id in multiTeachers:
                    multiTeacherSession = [
                        session for session in self.state if session.course.id == multiTeacherCourseId][0]
                    sessionsOfTeacher.append(multiTeacherSession)
                for day in range(5):
                    sessionsOfDay = list(filter(
                        lambda session: session.day == day, sessionsOfTeacher))
                    if len(sessionsOfDay) != 0:
                        ordered = sorted(
                            sessionsOfDay, key=lambda session: session.hour)
                        print(f'\n{getDayName(day)}\n')
                        for session in ordered:
                            print(
                                f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name} ({getDepartmentShortName(session.course.department)}-{session.course.year})')

    def printSemesterCollisions(self):
        print(
            f'\n----- Donem Cakismalari ({len(self.semesterCollisions) + len(self.languageSessionViolations)}) -----\n')
        for collision in self.semesterCollisions:
            for session in collision:
                print(
                    f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name}')
            print()

        for index, day in self.languageSessionViolations:
            print(
                f'{getDayName(day)}: 16.00 - 18.00 - Yabancı Dil ({index})')

    def printTeacherCollisions(self):
        print(
            f'\n----- Ogretmen Cakismalari ({len(self.teacherCollisions) + len(self.multiTeacherCollisions)}) -----\n')

        for collision in self.teacherCollisions:
            print(
                f'{collision[0].teacher.firstName} {collision[0].teacher.lastName}')
            for session in collision:
                print(
                    f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name} ({getDepartmentShortName(session.course.department)})')
            print()
        for collision in self.multiTeacherCollisions:
            print(
                f'{collision[0].teacher.firstName} {collision[0].teacher.lastName}')
            for session in collision:
                print(
                    f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name} ({getDepartmentShortName(session.course.department)})')
            print()

    def printBreakHourViolations(self):
        print(
            f'\n----- Öğle Arası İhlalleri ({len(self.fridayBreakViolations) + len(self.breakHourViolations) + len(self.departmentMeetingViolations)}) -----\n')
        for index, day in self.breakHourViolations:
            print(f'Öğle Arası İhlali: {index}, {getDayName(day)}')

        for index in self.fridayBreakViolations:
            print(f'Cuma Arası İhlali: {index}')

        for index in self.departmentMeetingViolations:
            print(f'Bölüm Toplantısı İhlali: {index}')

    def printFreeDays(self):
        print(f'\n----- Boş Günler ({len(self.freeDays)}) -----\n')
        for semester, day in self.freeDays:
            print(
                f'{getSemesterShortName(semester)}, {getDayName(day)}')

    def printAllSlotsUsedDays(self):
        print(
            f'\n----- Tüm Slotları Kullanılan Günler ({len(self.allSlotsUsedDays)}) -----\n')
        for semester, day in self.allSlotsUsedDays:
            print(
                f'{getSemesterShortName(semester)}, {getDayName(day)}')

    def printSingleSessionDays(self):
        print(
            f'\n----- Tek Seanslı Günler ({len(self.singleSessionDays)}) -----\n')
        for semester, day in self.singleSessionDays:
            print(
                f'{getSemesterShortName(semester)}, {getDayName(day)}')

    def printMultipleCourseSessions(self):
        print(
            f'\n----- Günde Birden Fazla Seanslı Dersler ({len(self.multipleCourseSessions)}) -----\n')
        for course in self.multipleCourseSessions:
            for session in course:
                print(
                    f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name} ({getSemesterShortName((4 * session.course.department + session.course.year) - 1)})')
            print()

    def printCannotCollideViolations(self):
        print(
            f'\n----- Çakışmaması Gereken Seans Çakışmaları ({len(self.cannotCollideViolations)}) -----\n')

        for collision in self.cannotCollideViolations:
            for session in collision:
                print(
                    f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name}')
            print()

    def printSlotSpan(self):
        print(
            f'\n----- Toplam Slot Açıklığı ({sum([sum(semesterSlotSpan) for semesterSlotSpan in self.slotSpan])})-----\n')

    def printHardConstraintViolations(self):
        print('\n\n---------- HARD CONSTRAINT VIOLATIONS ----------')
        self.printSemesterCollisions()
        self.printTeacherCollisions()
        self.printBreakHourViolations()
        self.printAllSlotsUsedDays()

    def printSoftContraints(self):
        print('\n\n---------- SOFT CONSTRAINT VIOLATIONS ----------')
        self.printAvailabilityCollisions()
        self.printFreeDays()
        self.printSingleSessionDays()
        self.printMultipleCourseSessions()
        self.printCannotCollideViolations()
        self.printSlotSpan()

    def printInfo(self):
        self.print()
        self.printTeacherSessions()
        self.printAvailableSlots()
        self.printHardConstraintViolations()
        self.printSoftContraints()
        self.printFitness()

    def printAvailabilityCollisions(self):
        print(
            f'\n\n----- Uygunluk Cakismalari ({len(self.teacherAvailabilityViolations)}) -----\n')
        for collision in self.teacherAvailabilityViolations:
            for session in collision:
                print(f'{session.teacher.firstName} {session.teacher.lastName}')
                print(
                    f'{getDayName(session.day)}: {session.hour}.00 - {session.hour + session.length}.00 - {session.name}')
            print()

    def printAvailableSlots(self):
        print('\n\n----- Uygun Slotlar -----\n')
        for index, semester in enumerate(self.availableSlots):
            print(f'\n\n{index}:', end=' ')
            for index, day in enumerate(semester):
                print(f'{getDayName(index)}: {day}', end=', ')

    def printFitness(self):
        print(
            f'fitness: {round(self.fitness, 2)} (isValid: {self.isValid})')

    def calculateSemesterCollisions(self):
        collisions = []
        for semester in self.semesters:
            for day in range(5):
                sessionsOfDay = list(filter(
                    lambda session: session.day == day, semester))
                usedSlots = []
                sessionIds = []
                for session in sessionsOfDay:
                    usedSlots.extend(
                        list(range(session.hour, session.hour + session.length)))
                    sessionIds.extend([session.id] * session.length)
                collisionSlots = [
                    item for item, count in Counter(usedSlots).items() if count > 1]
                for collisionSlot in collisionSlots:
                    indices = [index for index, slot in enumerate(
                        usedSlots) if slot == collisionSlot]
                    collisionSessions = []
                    for index in indices:
                        collisionSession = list(filter(
                            lambda session: session.id == sessionIds[index], sessionsOfDay))[0]
                        collisionSessions.append(collisionSession)
                    collisions.append(collisionSessions)
        return collisions

    def calculateTeacherCollisions(self):
        collisions = []
        for sessionsOfTeacher in self.teacherSessions:
            for day in range(5):
                sessionsOfDay = list(filter(
                    lambda session: session.day == day, sessionsOfTeacher))
                usedSlots = []
                sessionIds = []

                for session in sessionsOfDay:
                    usedSlots.extend(
                        list(range(session.hour, session.hour + session.length)))
                    sessionIds.extend([session.id] * session.length)
                collisionSlots = [
                    item for item, count in Counter(usedSlots).items() if count > 1]
                for collisionSlot in collisionSlots:
                    indices = [index for index, slot in enumerate(
                        usedSlots) if slot == collisionSlot]
                    collisionSessions = []
                    for index in indices:
                        collisionSession = list(filter(
                            lambda session: session.id == sessionIds[index], sessionsOfDay))[0]
                        collisionSessions.append(collisionSession)
                    collisions.append(collisionSessions)
        return collisions

    def calculateMultiTeacherSessionCollisions(self):
        collisions = []

        for teacherId in multiTeachers:
            sessionsOfTeacher = list(
                filter(lambda session: session.teacher.id == teacherId, self.state))
            multiTeacherSession = [
                session for session in self.state if session.course.id == multiTeacherCourseId][0]
            sessionsOfTeacher.append(multiTeacherSession)
            # for session in sessionsOfTeacher:
            #     print(f'{session.name}, {session.day}, {session.hour}')

            for day in range(5):
                sessionsOfDay = list(filter(
                    lambda session: session.day == day, sessionsOfTeacher))
                usedSlots = []
                sessionIds = []

                for session in sessionsOfDay:
                    usedSlots.extend(
                        list(range(session.hour, session.hour + session.length)))
                    sessionIds.extend([session.id] * session.length)
                collisionSlots = [
                    item for item, count in Counter(usedSlots).items() if count > 1]
                for collisionSlot in collisionSlots:
                    indices = [index for index, slot in enumerate(
                        usedSlots) if slot == collisionSlot]
                    collisionSessions = []
                    for index in indices:
                        collisionSession = list(filter(
                            lambda session: session.id == sessionIds[index], sessionsOfDay))[0]
                        collisionSessions.append(collisionSession)
                    collisions.append(collisionSessions)

                # if sum(collisionCount) > 0:
                #     print(collisionCount)
                #     print(sessionsOfTeacher[0].teacher.id)
        return collisions

    def calculateTeacherAvailabilityViolations(self):
        collisions = []
        for sessionsOfTeacher in self.teacherSessions:
            teacher = sessionsOfTeacher[0].teacher
            for day in range(5):
                sessionsOfDay = list(filter(
                    lambda session: session.day == day, sessionsOfTeacher))

                usedSlots = []
                sessionIds = []
                unavailableSlotsOfDay = teacher.unavailable[day]

                for session in sessionsOfDay:
                    usedSlots.extend(
                        list(range(session.hour, session.hour + session.length)))
                    sessionIds.extend([session.id] * session.length)

                collisionSlots = [
                    slot for slot in usedSlots if slot in unavailableSlotsOfDay]
                # collisionSlots = [
                #     item for item, count in Counter(usedSlots).items() if count > 1]
                for collisionSlot in collisionSlots:
                    indices = [index for index, slot in enumerate(
                        usedSlots) if slot == collisionSlot]
                    collisionSessions = []
                    for index in indices:
                        collisionSession = list(filter(
                            lambda session: session.id == sessionIds[index], sessionsOfDay))[0]
                        collisionSessions.append(collisionSession)
                    collisions.append(collisionSessions)
        return collisions

    def calculateCannotCollideViolations(self):
        violations = []
        for session in self.state:
            sessionSlots = list(
                range(session.hour, session.hour + session.length))
            sessionsToNotCollide = [
                s for s in self.state if s.course.id in session.course.cannotCollideWith]
            for sessionToNotCollide in sessionsToNotCollide:
                if session.day == sessionToNotCollide.day:
                    sessionToNotCollideSlots = list(range(
                        sessionToNotCollide.hour, sessionToNotCollide.hour + sessionToNotCollide.length))
                    sessionToNotCollideSlots.extend(sessionSlots)
                    collisionSlots = [
                        item for item, count in Counter(sessionToNotCollideSlots).items() if count > 1]
                    if collisionSlots and (sessionToNotCollide, session) not in violations:
                        violations.append((session, sessionToNotCollide))

        return violations

    def calculateHasAllSessions(self):
        return len(self.state) == 87

    def calculateAllSlotsUsedDays(self):
        allSlotsUsedDays = []
        for semesterIndex, semester in enumerate(self.availableSlots):
            for index, day in enumerate(semester):
                if not day and index in [0, 1, 3]:
                    allSlotsUsedDays.append((semesterIndex, index))
        return allSlotsUsedDays

    def calculateConstraints(self):
        breakViolations = []
        meetingViolations = []
        fridayViolations = []
        freeDays = []
        singleSessionDays = []
        multipleSessions = []
        totalSlotSpan = []
        languageSessionViolations = []
        allAvailableSlots = []

        for index, semester in enumerate(self.semesters):
            semesterSlotSpan = []
            semesterAvailableSlots = calculateAvailableSlots()
            for day in range(5):
                # sessionsOfDay = list(filter(
                #     lambda session: session.day == day, semester))
                sessionsOfDay = [
                    session for session in semester if session.day == day]
                usedSlots = []
                for session in sessionsOfDay:
                    usedSlots.extend(
                        list(range(session.hour, session.hour + session.length)))

                semesterAvailableSlots[day] = [
                    slot for slot in semesterAvailableSlots[day] if slot not in usedSlots]

                # Check for break violations:
                if day in [0, 1, 3] and (12 in usedSlots and 13 in usedSlots):
                    breakViolations.append((index, day))

                # Check for free days
                if len(usedSlots) == 0 and day not in [1, 2]:
                    freeDays.append((index, day))

                # Check for friday violations
                if day == 4 and (12 in usedSlots or 13 in usedSlots):
                    fridayViolations.append(index)

                # Check for meeting violations
                if day == 2 and 13 in usedSlots:
                    meetingViolations.append(index)

                # Check for language session violations
                if day in [1, 2] and (16 in usedSlots or 17 in usedSlots):
                    languageSessionViolations.append((index, day))

                # Check for single-session days
                if len(sessionsOfDay) == 1 and day in [0, 3, 4]:
                    singleSessionDays.append([index, day])
                elif len(sessionsOfDay) == 0 and day in [1, 2]:
                    singleSessionDays.append([index, day])

                # Check for multiple sessions of same course in the same day
                courseIds = [session.course.id for session in sessionsOfDay]
                multipleSessionCourseIds = [
                    item for item, count in Counter(courseIds).items() if count > 1]
                for multipleSessionCourseId in multipleSessionCourseIds:
                    sessionsOfCourse = list(filter(
                        lambda session: session.course.id == multipleSessionCourseId, sessionsOfDay))
                    multipleSessions.append(sessionsOfCourse)

                # Calculate the slot span
                if len(usedSlots) != 0:
                    earliestSlot = min(usedSlots)
                    latestSlot = max(usedSlots)
                    slotSpan = latestSlot - earliestSlot + 1
                    semesterSlotSpan.append(slotSpan)
                else:
                    semesterSlotSpan.append(0)

            totalSlotSpan.append(semesterSlotSpan)
            allAvailableSlots.append(semesterAvailableSlots)
        # print(totalSlotSpan)
        # print(sum([sum(semesterSlotSpan)
        #       for semesterSlotSpan in totalSlotSpan]))

        # for course in multipleSessions:
        #     for session in course:
        #         print(
        #             f'{session.hour}.00 - {session.hour + session.length}.00 - {session.name}')
        violations = [breakViolations, meetingViolations, fridayViolations, languageSessionViolations,
                      freeDays, singleSessionDays, multipleSessions, totalSlotSpan, allAvailableSlots]

        return violations

    def calculateAvailableSlots():
        allAvailableSlots = []
        for semester in self.semesters:
            for day in range(5):
                sessionsOfDay = list(filter(
                    lambda session: session.day == day, semester))
                usedSlots = []
                allSlots = [list(range(9, 18))] * 5
                for session in sessionsOfDay:
                    usedSlots.extend(
                        list(range(session.hour, session.hour + session.length)))

    def calculateFitness(self):
        # ! Hard Constraints
        semesterCollisionCount = len(self.semesterCollisions)
        languageSessionViolationCount = len(self.languageSessionViolations)
        teacherCollisionCount = len(self.teacherCollisions)
        multiTeacherCollisionCount = len(self.multiTeacherCollisions)
        breakHourViolationCount = len(self.breakHourViolations)
        fridayBreakViolationCount = len(self.fridayBreakViolations)
        departmentMeetingViolationCount = len(self.departmentMeetingViolations)
        allSlotsUsedDaysCount = len(self.allSlotsUsedDays)

        # ? Soft Constraints
        teacherAvailabilityViolationCount = len(
            self.teacherAvailabilityViolations)
        freeDayCount = len(self.freeDays)
        singleSessionDayCount = len(self.singleSessionDays)
        multipleCourseSessionCount = len(self.multipleCourseSessions)
        cannotCollideViolationCount = len(self.cannotCollideViolations)
        slotSpan = sum([sum(semesterSlotSpan)
                       for semesterSlotSpan in self.slotSpan])

        score = 50.0
        isValid = False

        score -= 2 * (semesterCollisionCount + teacherCollisionCount +
                      multiTeacherCollisionCount + languageSessionViolationCount)
        score -= 3 * (fridayBreakViolationCount +
                      breakHourViolationCount + departmentMeetingViolationCount)
        score -= 3 * allSlotsUsedDaysCount

        hardConstraintsTotal = semesterCollisionCount + teacherCollisionCount + multiTeacherCollisionCount + fridayBreakViolationCount + \
            breakHourViolationCount + departmentMeetingViolationCount + \
            languageSessionViolationCount + allSlotsUsedDaysCount

        score -= 0.6 * cannotCollideViolationCount
        score -= 0.5 * multipleCourseSessionCount
        score -= 0.4 * teacherAvailabilityViolationCount
        score -= 0.3 * singleSessionDayCount
        # Total session slots (210) + Total break slots (48)
        score -= 0.2 * (slotSpan - 258)
        score += 2 * freeDayCount

        if (hardConstraintsTotal):
            score -= 0.1 * score
            isValid = False
        else:
            score += 10
            isValid = True

        return isValid, score


def getSemesterName(department, year):
    departmentName = 'Bilgisayar Mühendisliği' if department == 0 else 'Endüstri Mühendisliği'
    return f"{departmentName} - {year}. Sınıf"


def getSemesterShortName(semester):
    department = 'BM' if semester // 4 == 0 else 'EM'
    year = semester % 4 + 1
    return f'{department}-{year}'


def getDepartmentName(department):
    return 'Bilgisayar Mühendisliği' if department == 0 else 'Endüstri Mühendisliği'


def getDepartmentShortName(department):
    return 'BM' if department == 0 else 'EM'


def getDayName(day):
    days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
    return days[day]


def generateInitialSemesterSchedule(semester):
    semester = copy.deepcopy(semester)
    available = copy.deepcopy(availableSlots)
    fixedSessions = [session for session in semester if session.isFixed]
    if fixedSessions:
        for fixedSession in fixedSessions:
            available[fixedSession.day] = [x for x in available[fixedSession.day]
                                           if x not in list(range(fixedSession.hour, fixedSession.hour + fixedSession.length))]
        # print(available)

    availableDays = [i for (i, x) in enumerate(available) if x]
    scheduledSessions = []
    shuffledSemester = random.sample(semester, len(semester))
    for session in semester:
        if session.isFixed:
            scheduledSessions.append(session)
            continue
        day, hour = selectAvailableDayAndHour(
            available, availableDays, session)
        if day is not False:
            session.day = day
            session.hour = hour
            available[day] = [x for x in available[day]
                              if x not in list(range(hour, hour + session.length))]
            availableDays = [i for (i, x) in enumerate(available) if x]

            scheduledSessions.append(session)

            # if 12 in available[day] or 13 in available[day]:
            #     scheduledSessions.append(session)
            # else:
            #     return False
        else:
            return False
    return scheduledSessions


def getSlotsOfFixedSessionsOfSemester(semester, fixedSessions):
    slots = [[], [], [], [], []]
    for fixedSession in fixedSessions:
        slots[fixedSession.day].extend(
            list(range(fixedSession.hour, fixedSession.hour + fixedSession.length)))
    return slots


def selectAvailableDayAndHour(available, availableDays, session):
    count = 0
    while True:
        day = random.choice(availableDays)
        hour = random.choice(available[day])
        slots = list(range(hour, hour + session.length))
        if 12 in slots and 13 in slots:
            count += 1
        else:
            isAllAvailable = all(x in available[day] for x in slots)
            if (isAllAvailable):
                return day, hour
            count += 1
        if count > 50:
            return False, False


def generateRandomSchedule():
    semesters = []
    for department in range(2):
        for year in range(1, 5):
            semester = list(filter(lambda session: session.course.department ==
                                   department and session.course.year == year, sessions))
            found = False
            while found == False or len(found) == 0:
                found = generateInitialSemesterSchedule(semester)
            semesters.extend(found)
    schedule = Schedule(semesters)
    return schedule


def generatePopulation(size):
    return [generateRandomSchedule() for _ in range(size)]


def importData():
    with open('./data/teachers.json', encoding='utf-8') as json_file:
        teachers_json = json.load(json_file)

    with open('./data/courses.json', encoding='utf-8') as json_file:
        courses_json = json.load(json_file)

    with open('./data/fixedSlots.json', encoding='utf-8') as json_file:
        fixedSlots = json.load(json_file)
    return teachers_json, courses_json, fixedSlots


def generateObjects():
    teachers = []
    for teacher in teachers_json:
        teachers.append(Teacher(
            teacher['id'], teacher['firstName'], teacher['lastName'], teacher['unavailable']))

    courses = []
    for course in courses_json:
        courses.append(Course(course['id'], course['name'],
                              course['code'], course['department'], course['year'], course['cannotCollideWith']))

    sessions = []
    index = 0
    for course in courses_json:
        for session in course['sessions']:
            sessions.append(
                Session(index, courses[course['id']], teachers[session['teacherId']],
                        session['length'], isLab=session.get("isLab", False), suffix=session.get("suffix", None), isFixed=session.get("isFixed", False), day=session.get("day", None), hour=session.get("hour", None)))
            index += 1

    return teachers, courses, sessions


def calculateAvailableSlots():
    allSlots = [list(range(9, 18))] * 5
    availableSlots = []
    for i in range(5):
        availableSlots.append(
            [x for x in allSlots[i] if x not in fixedSlots[i]])
    return availableSlots


def getMultiTeacherCourse():
    return next(
        (course['id'], course['teachers']) for course in courses_json if course.get("hasMultiTeachers", False) == True)


def performCrossover(schedule1, schedule2):
    index = random.choice(range(8))
    schedule1.semesters[index], schedule2.semesters[index] = schedule2.semesters[index], schedule1.semesters[index]

    newSchedule1 = Schedule(getStateFromSemesters(schedule1))
    newSchedule2 = Schedule(getStateFromSemesters(schedule2))

    return [newSchedule1, newSchedule2]


def getStateFromSemesters(schedule):
    sessions = []
    for semester in schedule.semesters:
        for session in semester:
            sessions.append(session)
    return sessions


def selection(population):
    scores = [
        # 10 * schedule.fitness if schedule.isValid else
        max(schedule.fitness, 0) for schedule in population]
    return random.choices(population, scores, k=SIZE)


def crossover(population):
    newPopulation = []
    for index in range(SIZE//2):
        newSchedules = performCrossover(
            population[index * 2], population[index * 2 + 1])
        newPopulation.extend(newSchedules)
    return newPopulation


def performMutation(schedule):
    mutation = random.choice(range(5))
    if mutation == 0:
        return mutateByMovingPeriod(schedule)
    if mutation == 1:
        return mutateBySwapingSessions(schedule)
    if mutation == 2:
        return mutateByMovingSession(schedule)
    if mutation == 3:
        return mutateBySwapingSessionsOfCourse(schedule)
    if mutation == 4:
        return mutateBySwapingSessionsThatCannotCollide(schedule)


def mutateByMovingPeriod(schedule):
    if schedule.breakHourViolations:
        semester, day = random.choice(schedule.breakHourViolations)
    else:
        # semester = random.randint(0, 7)
        # day = random.randint(0, 4)
        return schedule
    availableSlotsOfDay = schedule.availableSlots[semester][day]
    morningSlots = [
        slot for slot in availableSlotsOfDay if slot in [9, 10, 11]]
    eveningSlots = [
        slot for slot in availableSlotsOfDay if slot in [14, 15, 16, 17]]

    sessionsOfDay = list(filter(
        lambda session: session.day == day, schedule.semesters[semester]))
    startTime = [session.hour for session in sessionsOfDay]
    morningPossible = []
    eveningPossible = []
    if sum(morningSlots):
        morningPossible = [slot for slot in range(
            morningSlots[0], 13) if slot in startTime]
    if sum(eveningSlots):
        eveningPossible = [slot for slot in range(
            12, eveningSlots[-1] + 1) if slot in startTime]

    period = random.randint(0, 1)
    if period == 0 and morningPossible:
        toMutate = [
            session for session in sessionsOfDay if session.hour in morningPossible]

        # make sure that a fixed session won't mutate
        for session in toMutate:
            if session.isFixed:
                return schedule

        for session in toMutate:
            if session.hour > 9:
                session.hour -= 1
        return Schedule(schedule.state)

    if period == 1 and eveningPossible:
        toMutate = [
            session for session in sessionsOfDay if session.hour in eveningPossible]

        # make sure that a fixed session won't mutate
        for session in toMutate:
            if session.isFixed:
                return schedule

        for session in toMutate:
            if session.hour + session.length < 18:
                session.hour += 1
        return Schedule(schedule.state)

    return schedule


def mutateBySwapingSessions(schedule):
    if schedule.teacherCollisions:
        collision = random.choice(schedule.teacherCollisions)
    # elif schedule.cannotCollideViolations:
    #     collision = random.choice(schedule.cannotCollideViolations)
    else:
        return schedule

    for collidedSession in collision:
        swapableSessions = [session for session in schedule.state if isSafeToSwapTeacherCollision(
            session, collidedSession)]
        if swapableSessions:
            sessionToSwap = random.choice(swapableSessions)
            collidedSession.day, sessionToSwap.day = sessionToSwap.day, collidedSession.day
            collidedSession.hour, sessionToSwap.hour = sessionToSwap.hour, collidedSession.hour
            newSchedule = Schedule(schedule.state)
            # newSchedule.print()
            # newSchedule.printTeacherCollisions()
            return newSchedule

    return schedule


def mutateBySwapingSessionsThatCannotCollide(schedule):
    if schedule.cannotCollideViolations:
        collision = random.choice(schedule.cannotCollideViolations)
    else:
        return schedule

    for collidedSession in collision:
        swapableSessions = [session for session in schedule.state if isSafeToSwapTeacherCollision(
            session, collidedSession)]
        if swapableSessions:
            sessionToSwap = random.choice(swapableSessions)
            collidedSession.day, sessionToSwap.day = sessionToSwap.day, collidedSession.day
            collidedSession.hour, sessionToSwap.hour = sessionToSwap.hour, collidedSession.hour
            newSchedule = Schedule(schedule.state)
            # newSchedule.print()
            # newSchedule.printTeacherCollisions()
            return newSchedule

    return schedule


def mutateByRandomlySwapingSessions(schedule):
    for _ in range(10):
        chosenSession = random.choice(schedule.state)
        swapableSessions = [session for session in schedule.state if isSafeToSwapTeacherCollision(
            session, chosenSession)]
        if swapableSessions:
            sessionToSwap = random.choice(swapableSessions)
            chosenSession.day, sessionToSwap.day = sessionToSwap.day, chosenSession.day
            chosenSession.hour, sessionToSwap.hour = sessionToSwap.hour, chosenSession.hour
            newSchedule = Schedule(schedule.state)
            # newSchedule.print()
            # newSchedule.printTeacherCollisions()
            return newSchedule
    return schedule


def isSafeToSwapTeacherCollision(session, collidedSession):
    if session.course.id != collidedSession.course.id and \
            session.teacher.id != collidedSession.teacher.id and \
            session.course.year == collidedSession.course.year and \
            session.course.department == collidedSession.course.department and \
            session.length == collidedSession.length and \
            session.isFixed == False and \
            collidedSession.isFixed == False:
        return True
    else:
        return False


def mutateByMovingSession(schedule):
    # schedule.print()
    # schedule.printSingleSessionDays()

    if schedule.singleSessionDays:
        semesterIndex, day = random.choice(schedule.singleSessionDays)
    else:
        return schedule

    semester = schedule.semesters[semesterIndex]
    if day in [1, 2]:
        return schedule

    sessions = [session for session in semester if session.day == day]
    if len(sessions) == 0:
        return schedule
    session = sessions[0]

    if session.isFixed:
        return schedule

    availableSlots = schedule.availableSlots[semesterIndex]

    length = session.length
    possibleSlots = []
    if length == 2:
        for availableSlotsOfDay in availableSlots:
            possibleSlots.append([
                slot for slot in availableSlotsOfDay if slot + 1 in availableSlotsOfDay])
    if length == 3:
        for availableSlotsOfDay in availableSlots:
            possibleSlots.append([slot for slot in availableSlotsOfDay if slot +
                                 1 in availableSlotsOfDay and slot + 2 in availableSlotsOfDay])
    allPossibleDays = []
    for day, daySlots in enumerate(possibleSlots):
        if daySlots:
            allPossibleDays.append(day)

    chosenDay = random.choice(allPossibleDays)
    chosenSlot = random.choice(possibleSlots[chosenDay])

    session.day = chosenDay
    session.hour = chosenSlot

    newSchedule = Schedule(schedule.state)
    # newSchedule.print()
    # newSchedule.printSingleSessionDays()

    return newSchedule


def mutateBySwapingSessionsOfCourse(schedule):
    if schedule.multipleCourseSessions:
        sessionsOfCourse = random.choice(schedule.multipleCourseSessions)
        # for session in sessionsOfCourse:
        #     session.printSchedule()
        # schedule.print()
        # schedule.printMultipleCourseSessions()
    else:
        return schedule

    for sessionOfCourse in sessionsOfCourse:
        swapableSessions = [session for session in schedule.state if isSafeToSwapMultipleSessions(
            session, sessionOfCourse)]
        if swapableSessions:
            sessionToSwap = random.choice(swapableSessions)
            sessionOfCourse.day, sessionToSwap.day = sessionToSwap.day, sessionOfCourse.day
            sessionOfCourse.hour, sessionToSwap.hour = sessionToSwap.hour, sessionOfCourse.hour
            newSchedule = Schedule(schedule.state)
            # newSchedule.print()
            # newSchedule.printMultipleCourseSessions()
            return newSchedule

    return schedule


def isSafeToSwapMultipleSessions(session, sessionOfCourse):
    if session.course.id != sessionOfCourse.course.id and \
            session.course.year == sessionOfCourse.course.year and \
            session.course.department == sessionOfCourse.course.department and \
            session.length == sessionOfCourse.length and \
            session.isFixed == False and \
            sessionOfCourse.isFixed == False:
        return True
    else:
        return False


def mutation(population):
    newPopulation = []
    for schedule in population:
        willMutate = random.choices([True, False], [0.1, 0.9], k=1)
        if willMutate:
            newPopulation.append(performMutation(schedule))
        else:
            newPopulation.append(schedule)
    return newPopulation


def exportSchedule(schedule, name='latest'):
    dbfile = open(f'output/{name}', 'ab')
    pickle.dump(schedule, dbfile)
    dbfile.close()


def importSchedule(name='latest', info=False):
    dbfile = open(f'output/{name}', 'rb')
    schedule = pickle.load(dbfile)
    dbfile.close()
    if info:
        schedule.printInfo()
    return schedule


def saveToExcel(schedule):
    wb = openpyxl.load_workbook("template.xlsx")
    sheets = wb.sheetnames

    for index, sheet in enumerate(sheets):
        ws = wb[sheet]

        semester = schedule.semesters[index]
        for day in range(5):
            dayColumn = getColumnOfDay(day)
            sessionsOfDay = [
                session for session in semester if session.day == day]
            sessions = sorted(
                sessionsOfDay, key=lambda session: session.hour)
            for session in sessions:
                slots = list(
                    range(session.hour, session.hour + session.length))
                mergeString = f"{dayColumn}{slots[0]}:{dayColumn}{slots[-1]}"
                mergeStart = f"{dayColumn}{slots[0]}"
                ws.merge_cells(mergeString)
                cellString = f"{session.course.code}\n{session.name}\n{session.teacher.firstName} {session.teacher.lastName}"
                ws[mergeStart] = cellString
                ws[mergeStart].font = openpyxl.styles.Font(
                    bold=True, name='TimesNewRoman', size='11')
                ws[mergeStart].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                                     wrapText=True)

            if day in [1, 2]:
                ws.merge_cells(f"{dayColumn}16:{dayColumn}17")
                mergeStart = f"{dayColumn}16"
                ws[mergeStart].font = openpyxl.styles.Font(
                    bold=True, name='TimesNewRoman', size='11')
                ws[mergeStart] = 'Yabancı Dil'

        setBorder(ws)

    wb.save(f"./export/{round(schedule.fitness, 2)}.xlsx")
    cwd = os.getcwd()
    os.startfile(f"{cwd}\\export\\{round(schedule.fitness, 2)}.xlsx")


def setBorder(ws):
    thick = openpyxl.styles.Side(border_style="thick", color="000000")

    for row in ws['B18:G18']:
        for cell in row:
            cell.border = openpyxl.styles.Border(
                top=thick)


def getColumnOfDay(day):
    if day == 0:
        return 'C'
    if day == 1:
        return 'D'
    if day == 2:
        return 'E'
    if day == 3:
        return 'F'
    if day == 4:
        return 'G'


def printInitilaPopulationFitness(population):
    sortedPopulation = sorted(
        population, key=lambda schedule: schedule.fitness, reverse=True)
    print(f'\n INITIAL POPULATION')
    for schedule in sortedPopulation:
        schedule.printFitness()
    average = sum([schedule.fitness for schedule in sortedPopulation]
                  ) / len(sortedPopulation)
    print(f'Average: {average}')


def printPopulationFitness(population, generation, stagnation, bestSoFar, showAll=False):

    print(f'\nGENERATION {generation + 1} (stagnation= {stagnation})')
    if showAll:
        for schedule in population:
            schedule.printFitness()

    average = sum([schedule.fitness for schedule in population]
                  ) / len(population)
    print(f'Average: {round(average, 2)}')
    print(f'Best of Generation: {round(population[0].fitness, 2)}')
    print(f'Best So Far: {round(bestSoFar.fitness, 2)}', end=' ')
    print('VALID') if bestSoFar.isValid else print('INVALID')


def evolution(size, limit, population):
    bestScore = 0
    bestSoFar = None
    stagnation = 0
    generation = 0

    printInitilaPopulationFitness(population)
    mutateBySwapingSessionsOfCourse(population[0])

    while stagnation <= limit:
        population = selection(population)
        population = crossover(population)
        population = mutation(population)

        sortedPopulation = sorted(
            population, key=lambda schedule: schedule.fitness, reverse=True)

        bestOfGeneration = sortedPopulation[0]

        if bestOfGeneration.fitness > bestScore:
            bestScore = bestOfGeneration.fitness
            bestSoFar = bestOfGeneration
            stagnation = 0

        printPopulationFitness(
            sortedPopulation, generation, stagnation, bestSoFar)

        stagnation += 1
        generation += 1

    exportSchedule(
        bestSoFar, name=f'{round(bestSoFar.fitness, 2)}, {limit}, {size}')
    bestSoFar.printInfo()

    return bestSoFar


SIZE = 64
LIMIT = 25

teachers_json, courses_json, fixedSlots = importData()
teachers, courses, sessions = generateObjects()
multiTeacherCourseId, multiTeachers = getMultiTeacherCourse()
availableSlots = calculateAvailableSlots()
population = generatePopulation(SIZE)
best = evolution(SIZE, LIMIT, population)
saveToExcel(best)
# imported = importSchedule(name='Problem2', info=True)
