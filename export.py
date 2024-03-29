import pickle
import openpyxl
import os
from utils import languageSlots


def exportSchedule(schedule, name='latest'):
    dbfile = open(f'output/{name}', 'ab')
    pickle.dump(schedule, dbfile)
    dbfile.close()


def importSchedule(name='latest', info=False):
    dbfile = open(f'output/{name}', 'rb')
    schedule = pickle.load(dbfile)
    dbfile.close()
    if info:
        schedule.printInfo()
    return schedule


def saveToExcel(schedule, openFile=False, filename=None):
    wb = openpyxl.load_workbook("template.xlsx")
    sheets = wb.sheetnames

    for index, sheet in enumerate(sheets):
        ws = wb[sheet]

        semester = schedule.semesters[index]
        for day in range(5):
            dayColumn = getColumnOfDay(day)
            sessionsOfDay = [
                session for session in semester if session.day == day]
            sessions = sorted(
                sessionsOfDay, key=lambda session: session.hour)
            for session in sessions:
                slots = list(
                    range(session.hour, session.hour + session.length))
                mergeString = f"{dayColumn}{slots[0]}:{dayColumn}{slots[-1]}"
                mergeStart = f"{dayColumn}{slots[0]}"
                ws.merge_cells(mergeString)
                cellString = f"{session.course.code}\n{session.name}\n{session.teacher.firstName} {session.teacher.lastName}"
                ws[mergeStart] = cellString
                ws[mergeStart].font = openpyxl.styles.Font(
                    bold=True, name='TimesNewRoman', size='11')
                ws[mergeStart].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                                     wrapText=True)

            if day in languageSlots[0]:
                ws.merge_cells(
                    f"{dayColumn}{languageSlots[1][0]}:{dayColumn}{languageSlots[1][1]}")
                mergeStart = f"{dayColumn}{languageSlots[1][0]}"
                ws[mergeStart].font = openpyxl.styles.Font(
                    bold=True, name='TimesNewRoman', size='11')
                ws[mergeStart] = 'Yabancı Dil'

        setBorder(ws)
    if not filename:
        name = f"./excel/{round(schedule.fitness, 2)}.xlsx"
    else:
        name = f"./excel/{filename}.xlsx"
    wb.save(name)
    if openFile:
        cwd = os.getcwd()
        os.startfile(f"{cwd}\\excel\\{round(schedule.fitness, 2)}.xlsx")


def setBorder(ws):
    thick = openpyxl.styles.Side(border_style="thick", color="000000")

    for row in ws['B18:G18']:
        for cell in row:
            cell.border = openpyxl.styles.Border(
                top=thick)


def getColumnOfDay(day):
    if day == 0:
        return 'C'
    if day == 1:
        return 'D'
    if day == 2:
        return 'E'
    if day == 3:
        return 'F'
    if day == 4:
        return 'G'
